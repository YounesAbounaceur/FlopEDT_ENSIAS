# coding: utf-8
# !/usr/bin/python

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

from openpyxl import load_workbook

from base.models import Group, Module, Period, CourseType

from copy import copy

def as_text(value):
    if value is None:
        return ""
    return str(value)


def append_row(work_sheet, rows_to_append, row_number, rank, untill):
    row = rows_to_append[row_number-1]
    for cell in row[:untill]:
        new_cell = work_sheet.cell(row=rank, column=cell.col_idx, value= cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

empty_bookname='misc/deploy_database/empty_planif_file.xlsx'

def make_planif_file(department, empty_bookname=empty_bookname):
    new_book = load_workbook(filename=empty_bookname)
    empty_rows = list(new_book['empty'].rows)
    # We go through each period and create a sheet for each period
    for p in Period.objects.filter(department=department):
        new_book.create_sheet(p.name)
        sheet = new_book[p.name]
	
        ################ Writing line 1 with weeks ################
        rank = 1
        WEEK_COL = 7

        if p.starting_week < p.ending_week:
            weeks = p.ending_week - p.starting_week + 1
            cols = weeks + 7
            append_row(sheet, empty_rows, 1, rank, cols)
            for i in range(p.starting_week, p.ending_week+1):

                sheet.cell(row=rank, column=WEEK_COL).value = i
                WEEK_COL += 1

        else:
            weeks = (53 - p.starting_week) + p.ending_week
            cols = weeks + 7
            append_row(sheet, empty_rows, 1, rank, cols)
            for i in range(p.starting_week, 53):
                sheet.cell(row=rank, column=WEEK_COL).value = i
                WEEK_COL += 1

            for i in range(1, p.ending_week+1):

                sheet.cell(row=rank, column=WEEK_COL).value = i
                WEEK_COL += 1
        rank += 1
        append_row(sheet, empty_rows, 4, rank, cols)


        ################ A line per module per CourseType ################
        for mod in Module.objects.filter(period=p):
            for ct in CourseType.objects.filter(department=department):
                rank += 1
                append_row(sheet, empty_rows, 2, rank, cols)
                sheet.cell(row=rank, column=1).value = mod.abbrev
                sheet.cell(row=rank, column=2).value = ct.name
                sheet.cell(row=rank, column=3).value = 'Prof'
                sheet.cell(row=rank, column=4).value = 'Salle'
                sheet.cell(row=rank, column=5).value = 'Durée'
                sheet.cell(row=rank, column=6).value = 'Groupes'

                if Group.objects.filter(type__in=ct.group_types.all(), train_prog=mod.train_prog).count() > 0:
                    for g in Group.objects.filter(type__in=ct.group_types.all(), train_prog=mod.train_prog):
                        rank += 1
                        append_row(sheet, empty_rows, 3, rank, cols)
                        sheet.cell(row=rank, column=1).value = mod.abbrev
                        sheet.cell(row=rank, column=2).value = ct.name
                        sheet.cell(row=rank, column=6).value = g.nom
                else:
                    rank += 1
                    append_row(sheet, empty_rows, 3, rank, cols)
                    sheet.cell(row=rank, column=1).value = mod.abbrev
                    sheet.cell(row=rank, column=2).value = ct.name

        ################ Separating each course with a black line ################
            rank += 1
            append_row(sheet, empty_rows, 4, rank, cols)
        
        ############ TOTAL line ############
        rank += 1
        append_row(sheet, empty_rows, 5, rank, cols)
    
    ############ Adapting column widths ############
        for col in sheet.columns:
            length = len(as_text(col[0].value))
            adjusted_length = (length + 2) * 1.2
            sheet.column_dimensions[col[0].column].width = adjusted_length


    new_book.remove(new_book['empty'])
    filename='misc/deploy_database/planif_file_'+department.abbrev+'.xlsx'
    new_book.save(filename=filename)
