# coding: utf-8
# !/usr/bin/python

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
# 
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
# 
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
# 
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
# 
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

import os
import sys

sys.path.append("..")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "FlOpEDT.settings.local")

### Have to do this for it to work in 1.9.x!
from django.core.wsgi import get_wsgi_application

application = get_wsgi_application()
#############


from openpyxl import *
from base.models import Group, Module, Course, Room, CourseType, RoomType, TrainingProgramme, Dependency, Period, Department

from people.models import Tutor

from misc.assign_module_color import assign_color


def ReadPlanifWeek(department, book, feuille, semaine, an):
    sheet = book[feuille]
    period=Period.objects.get(name=feuille, department=department)

    # lookup week row:

    # lookup week column
    wc = 1
    for wr in [1, 5]:
        while wc < 50:
            wc += 1
            sem = sheet.cell(row=wr, column=wc).value
            if sem == float(semaine):
                WEEK_COL = wc
                break
    print("Semaine %s de %s : colonne %g" % (semaine, feuille, WEEK_COL))

    row = 2
    module_COL = 1
    nature_COL = 2
    prof_COL = 3
    salle_COL = 4
    groupe_COL = 6
    durée_COL = 5
    sumtotal = 0
    while 1:
        row += 1
        salle = sheet.cell(row=row, column=salle_COL).value
        module = sheet.cell(row=row, column=module_COL).value
        if salle is None:
            continue
        if salle.startswith("TOTAL"):
            # print "Sem %g de %s - TOTAL: %g"%(semaine, feuille,sumtotal)
            break

        Cell = sheet.cell(row=row, column=WEEK_COL)
        N = Cell.value
        if N is None:
            continue

        try:
            N = float(N)
            # handle dark green lines - Vert fonce
            assert isinstance(salle, str)
            if salle == "Salle":
                nominal = int(N)
                if N != nominal:
                    print('Valeur decimale ligne %g de %s, semaine %g : on la met a 1 !' % (row, feuille, semaine))
                    nominal = 1
                    # le nominal est le nombre de cours par groupe (de TP ou TD)
                if Cell.comment:
                    comments = Cell.comment.text.replace(' ', '').replace('\n', '').replace(',', ';').split(';')
                else:
                    comments = []

                sumtotal += nominal
                continue
            try:
                comments = comments
            except:
                comments = []
            # handle light green lines - Vert clair
            MODULE = Module.objects.get(abbrev=module, period=period)
            PROMO = MODULE.train_prog
            nature = sheet.cell(row=row, column=nature_COL).value
            salle = sheet.cell(row=row, column=salle_COL).value
            prof = sheet.cell(row=row, column=prof_COL).value
            grps = sheet.cell(row=row, column=groupe_COL).value
            COURSE_TYPE = CourseType.objects.get(name=nature, department=department)
            ROOMTYPE = RoomType.objects.get(name=salle, department=department)
            duration = int(duration)
            if prof is None:
                TUTOR, created = Tutor.objects.get_or_create(username='---')
                if created:
                    TUTOR.save()
                supp_profs=[]
            else:
                assert isinstance(prof, str) and prof is not None
                profs = prof.split(";")
                prof = profs[0]
                TUTOR = Tutor.objects.get(username=prof)
                supp_profs = profs[1:]
            SUPP_TUTORS = Tutor.objects.filter(username__in=supp_profs)

            if Cell.comment:
                local_comments = Cell.comment.text.replace(' ', '').replace('\n', '').replace(',', ';').split(';')
            else:
                local_comments = []

            if isinstance(grps, int) or isinstance(grps, float):
                grps = str(int(grps))
            if not grps:
                grps = []
            else:
                grps = grps.replace(' ', '').replace(',', ';').split(';')
            groupes = [str(g) for g in grps]

            GROUPS = list(Group.objects.filter(nom__in=groupes, train_prog=PROMO))

            N=int(N)
            Diff = N - len(groupes) * nominal
            if Diff != 0:
                print("Nombre incohérent ligne %g semaine %s de %s : %s \n" % (row, semaine, feuille, module))

            for i in range(N):
                GROUPE = GROUPS[i % len(groupes)]
                C = Course(tutor=TUTOR, type=COURSE_TYPE, module=MODULE, groupe=GROUPE, semaine=semaine, an=an,
                           room_type=ROOMTYPE, duration=duration)
                C.save()
                for sp in SUPP_TUTORS:
                    C.supp_tutor.add(sp)
                    C.save()
                for after_type in [x for x in comments + local_comments if x[0] == 'A']:
                    try:
                        n = int(after_type[1])
                        s = 2
                    except ValueError:
                        n = 1
                        s = 1
                    course_type = after_type[s:]
                    courses = Course.objects.filter(type__name=course_type, module=MODULE, semaine=semaine, an=an,
                                                   groupe__in = list(GROUPE.ancestor_groups()) + [GROUPE])
                    for course in courses[:n]:
                        P = Dependency(cours1=course, cours2=C)
                        P.save()

            if 'D' in comments or 'D' in local_comments  and N >= 2:
                for GROUPE in GROUPS:
                    Cours = Course.objects.filter(type=COURSE_TYPE, module=MODULE, groupe=GROUPE, an=an,
                                                  semaine=semaine)
                    for i in range(N//2):
                        P = Dependency(cours1=Cours[2*i], cours2=Cours[2*i+1], successifs=True)
                        P.save()
            if 'ND' in comments or 'ND' in local_comments  and N >= 2:
                for GROUPE in GROUPS:
                    Cours = Course.objects.filter(type=COURSE_TYPE, module=MODULE, groupe=GROUPE, an=an,
                                                  semaine=semaine)
                    P = Dependency(cours1=Cours[0], cours2=Cours[1], ND=True)
                    P.save()
        except Exception as e:
            print("Exception ligne %g semaine %s de %s : %s \n" % (row, semaine, feuille, module), e)
            raise


def extract_period(department, book, period):
    if period.starting_week < period.ending_week:
        for week in range(period.starting_week, period.ending_week + 1):
            ReadPlanifWeek(department, book, period.name, week, 2018)
    else:
        for week in range(period.starting_week, 53):
            ReadPlanifWeek(department, book, period.name, week, 2018)
        for week in range(1, period.ending_week + 1):
            ReadPlanifWeek(department, book, period.name, week, 2019)

def extract_planif(department, bookname=None):
    if bookname is None:
        bookname = 'misc/deploy_database/planif_file_'+department.abbrev+'.xlsx'
    book = load_workbook(filename=bookname, data_only=True)
    for period in Period.objects.filter(department=department):
        extract_period(department, book, period)
    assign_color(department)

