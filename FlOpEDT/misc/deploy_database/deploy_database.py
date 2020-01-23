# coding: utf-8
# !/usr/bin/python

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

import string, logging

from django.db import transaction
from openpyxl import load_workbook

from django.db.models import Q
from random import choice

from base.models import Room, RoomType, RoomGroup, TrainingProgramme, TrainingProgrammeDisplay,\
    Group, Module, GroupType, Period, Time, Day, Slot, CourseType, EdtVersion, UserPreference,\
    CoursePreference, Department

from base.weeks import annee_courante

from misc.check_slots import assign_day_time_numbers

from misc.generate_static_files import generate_group_file

from people.models import FullStaff, SupplyStaff, Tutor

from django.db import IntegrityError


bookname='misc/deploy_database/database_file.xlsx'
logger = logging.getLogger('base')

@transaction.atomic
def extract_database_file(bookname=bookname, department_name=None, department_abbrev=None):

    # Test department existence
    department, created = Department.objects.get_or_create(name=department_name, abbrev=department_abbrev)
    if not created:
        print(f"Department with abbrev {department_abbrev} already exists.")
        return

    book = load_workbook(filename=bookname, data_only=True)
    tutors_extract(department, book)
    rooms_extract(department, book)
    groups_extract(department, book)
    modules_extract(department, book)
    slots_extract(department, book)
    courses_extract(department, book)


def tutors_extract(department, book):

    sheet = book["Intervenants"]

    INTER_ID_ROW = 3
    id = sheet.cell(row=INTER_ID_ROW, column=1).value

    while id is not None :

        name = sheet.cell(row=INTER_ID_ROW, column=2).value
        last_name = sheet.cell(row=INTER_ID_ROW, column=3).value
        status = sheet.cell(row=INTER_ID_ROW, column=4).value
        email = sheet.cell(row=INTER_ID_ROW, column=5).value
        
        try:
            tutor = Tutor.objects.get(username=id)
            logger.debug(f'update tutor : [{id}]')
        except Tutor.DoesNotExist:

            try:
                params = { 'username': id, 'first_name': name, 'last_name': last_name, 'email': email, }    
                
                if status == "Permanent":
                    tutor = FullStaff(**params)
                else:
                    employer = sheet.cell(row=INTER_ID_ROW, column=9).value
                    position = sheet.cell(row=INTER_ID_ROW, column=8).value

                    params.update({'employer': employer, 'position': position})
                    tutor = SupplyStaff(**params)


                tutor.set_password("passe")
                tutor.is_tutor = True
                tutor.save()


            except IntegrityError as ie :
                print("A constraint has not been respected creation the Professor : \n", ie)
                pass
            else:
                logger.debug(f'create tutor with id:{id}')
        else:
            tutor.departments.add(department)

        INTER_ID_ROW += 1
        id = sheet.cell(row=INTER_ID_ROW, column=1).value

    print("Tutors extraction done")


def rooms_extract(department, book):

    sheet = book['Salles']
    ######################## Creating RoomTypes ####################################

    ROOM_CATEGORY_START_ROW = 20
    ROOM_CATEGORY_START_COL = 5
    
    row = ROOM_CATEGORY_START_ROW
    col = ROOM_CATEGORY_START_COL
    idCat = sheet.cell(row=row, column=col).value

    # Create temporary RoomType for import purposes. This type 
    # will be deleted at the end of the process
    temporay_room_random_key = ''.join(choice(string.ascii_lowercase + string.digits) for _ in range(6))
    temporay_room_type = RoomType.objects.create(department=department, name=f"temp_{department.abbrev}_{temporay_room_random_key}")

    while idCat is not None :
        try:
            RoomType.objects.get_or_create(department=department, name=idCat)
        except IntegrityError as ie:
            print("A constraint has not been respected creating the RoomType %s : \n" %idCat, ie)

        row += 1
        idCat = sheet.cell(row=row, column=col).value

    ######################## Creating Rooms ####################################

    ROOM_DECLARATION_START_ROW = 3
    ROOM_DECLARATION_COL = 1
    
    row = ROOM_DECLARATION_START_ROW
    col = ROOM_DECLARATION_COL
    idRoom = sheet.cell(row=row, column=col).value

    while idRoom is not None :

        try:
            room, _ = Room.objects.get_or_create(name=idRoom)
            
            room_group, _ = RoomGroup.objects.get_or_create(name=idRoom)
            room_group.types.add(temporay_room_type)
            
            # Ensure that a room_group exits with the same roomid
            room.subroom_of.add(room_group)


        except IntegrityError as ie:
            print("A constraint has not been respected creating the Room %s : \n" %idRoom, ie)

        row += 1
        idRoom = sheet.cell(row=row, column=col).value

    ######################## Creating RoomGroups ####################################

    ROOMGROUP_DECLARATION_START_ROW = 3
    ROOMGROUP_DECLARATION_COL = 2

    row = ROOMGROUP_DECLARATION_START_ROW
    col = ROOMGROUP_DECLARATION_COL
    room_group_id = sheet.cell(row=row, column=col).value

    while room_group_id is not None :

        try:
            if not Room.objects.filter(name=room_group_id).exists():
                room_group = RoomGroup.objects.create(name=room_group_id)
                room_group.types.add(temporay_room_type)
            else:
                print(f"A custom group can't have the same name thant an existing Room : {room_group_id}")

        except IntegrityError as ie:
            print("A constraint has not been respected creating the RoomGroup %s : \n" %room_group_id, ie)

        row += 1
        room_group_id = sheet.cell(row=row, column=col).value

    ######################## Filling the RoomGroups with Rooms ####################################

    ROOMGROUP_DEFINITION_START_ROW = 4
    ROOMGROUP_DEFINITION_START_COL = 5

    row = ROOMGROUP_DEFINITION_START_ROW
    col = ROOMGROUP_DEFINITION_START_COL
    idGroup = sheet.cell(row=row, column=col).value

    while idGroup is not None :

        col = ROOMGROUP_DEFINITION_START_COL + 1
        idRoom = sheet.cell(row=row, column=col).value

        while idRoom is not None :

            logger.info(f"Add room [{idRoom}] to group : {idGroup}")
            
            try:                
                room = Room.objects.get(name=idRoom)
                room_group = RoomGroup.objects.get(name=idGroup, types__in=[temporay_room_type,])
                room.subroom_of.add(room_group)

            except Room.DoesNotExist:
                print(f"unable to find room '{idRoom}' with correct RoomType'")
            
            except RoomGroup.DoesNotExist:
                print(f"unable to find  RoomGroup '{idGroup}' with correct RoomType'")                            

            col += 1
            idRoom = sheet.cell(row=row, column=col).value

        row += 1
        idGroup = sheet.cell(row=row,
                             column=ROOMGROUP_DEFINITION_START_COL).value

    ######################## Giving a RoomType to each RoomGroup ####################################

    row = ROOM_CATEGORY_START_ROW
    col = ROOM_CATEGORY_START_COL
    idCat = sheet.cell(row=row, column=col).value

    while idCat is not None :

        col = ROOM_CATEGORY_START_COL + 1
        room_group_id = sheet.cell(row=row, column=col).value

        room_type = RoomType.objects.get(department=department, name=idCat)

        while room_group_id is not None :
            try:
                # Test if group is a common room based group or a department custom group
                try:
                    room_group = RoomGroup.objects.get(subrooms__id=room_group_id)
                except:
                    room_group = RoomGroup.objects.get(name=room_group_id, types__in=[temporay_room_type,])

                room_group.types.add(room_type)
            except RoomGroup.DoesNotExist:
                print(f"unable to find  RoomGroup '{idroom_group_idGroup}'")

            col += 1
            room_group_id = sheet.cell(row=row, column=col).value

        row += 1
        idCat = sheet.cell(row=row, column=ROOM_CATEGORY_START_COL).value

    temporay_room_type.delete()
    print("Rooms extraction done")


# groups_extract
# Creates the groups, the training programs, and
# and fills the groups with their parent groups

def groups_extract(department, book):

    sheet = book['Groupes']

    ######################## Creating the TrainingPrograms ####################################

    TP_ROW = 3
    TP_COL = 13

    idTP = sheet.cell(row=TP_ROW, column=TP_COL).value

    while idTP is not None:

        verif = TrainingProgramme.objects.filter(abbrev=idTP)

        if not verif.exists():

            nameTP = sheet.cell(row=TP_ROW, column=TP_COL + 1).value

            try:
                trainingProg = TrainingProgramme(department=department, name=nameTP, abbrev=idTP)
                trainingProg.save()
            except IntegrityError as ie:
                print("A constraint has not been respected creating the TrainingProgramme %s : \n" % idTP, ie)
                pass

        TP_ROW += 1
        idTP = sheet.cell(row=TP_ROW, column=TP_COL).value

    ######################## Creating the GroupTypes ####################################

    GT_ROW = 17

    idGroupType = sheet.cell(row=GT_ROW, column=13).value

    while idGroupType is not None:

        verif = GroupType.objects.filter(name=idGroupType, department=department)

        if not verif.exists():

            try:

                gt = GroupType(name=idGroupType, department=department)
                gt.save()

            except IntegrityError as ie:
                print("A constraint has not been respected creating the GroupType %s : \n" % idGroupType, ie)
                pass

        GT_ROW += 1
        idGroupType = sheet.cell(row=GT_ROW, column=13).value


    ######################## Creating the Groups ####################################

    GROUP_ROW = 3

    idGroup = sheet.cell(row=GROUP_ROW, column=1).value

    while idGroup is not None:

        tpGr = sheet.cell(row=GROUP_ROW, column=2).value
        verif = Group.objects.filter(nom=idGroup, train_prog__abbrev=tpGr)

        if not verif.exists():

            try:

                tpGr = sheet.cell(row=GROUP_ROW, column=2).value
                tpGroup = TrainingProgramme.objects.get(abbrev=tpGr)

                gt = sheet.cell(row=GROUP_ROW, column=5).value
                groupType = GroupType.objects.get(name=gt, department=department)

                group = Group(nom=idGroup, size=0, train_prog=tpGroup, type=groupType)
                group.save()

            except IntegrityError as ie:
                print("A constraint has not been respected creating the Group %s : \n" % idGroup, ie)
                pass

        GROUP_ROW += 1
        idGroup = sheet.cell(row=GROUP_ROW, column=1).value


    ######################## Filling the Groups with their parent groups ####################################

    GROUP_ROW = 3

    idGroup = sheet.cell(row=GROUP_ROW, column=1).value

    while idGroup is not None:

        tpGr = sheet.cell(row=GROUP_ROW, column=2).value
        p_group = sheet.cell(row=GROUP_ROW, column=3).value
        p_group2 = sheet.cell(row=GROUP_ROW, column=4).value

        if p_group is not None:

            parent_group = Group.objects.get(nom=p_group, train_prog__abbrev=tpGr)

            group = Group.objects.get(nom=idGroup, train_prog__abbrev=tpGr)

            group.parent_groups.add(parent_group)

            group.save()

            if p_group2 is not None:

                parent_group = Group.objects.get(nom=p_group2, train_prog__abbrev=tpGr)

                group = Group.objects.get(nom=idGroup, train_prog__abbrev=tpGr)

                group.parent_groups.add(parent_group)

                group.save()

        GROUP_ROW += 1
        idGroup = sheet.cell(row=GROUP_ROW, column=1).value

######################## Defining basic groups ####################################

    for g in Group.objects.all():

        isbasic = True

        for g1 in Group.objects.all():

            if g in g1.parent_groups.all():

                isbasic = False

        g.basic = isbasic
        g.save()

######################## Defining Periods ####################################

    PERIOD_ROW = 12

    id_per = sheet.cell(row=PERIOD_ROW, column=7).value

    while id_per is not None:

        verif = Period.objects.filter(department=department, name = id_per)

        if not verif.exists():

            s_week = int(sheet.cell(row=PERIOD_ROW, column=8).value)
            e_week = int(sheet.cell(row=PERIOD_ROW, column=9).value)

            try:

                period = Period.objects.create(
                                            name=id_per,
                                            department=department,
                                            starting_week=s_week,
                                            ending_week=e_week)

            except IntegrityError as ie:
                print("A constraint has not been respected creating the Period %s : \n" % id_per, ie)
                pass

        PERIOD_ROW += 1
        id_per = sheet.cell(row=PERIOD_ROW, column=7).value

    for index, tp in enumerate(TrainingProgramme.objects.filter(department=department)):        
        TrainingProgrammeDisplay.objects.get_or_create(training_programme=tp, row=index)
        

    #generate_group_file(department.abbrev)

    print("Groups extraction done")


def modules_extract(department, book):

    sheet = book["Modules"]

    MODULE_ROW=3

    idMod=sheet.cell(row=MODULE_ROW, column=1).value


    while idMod is not None:

        tpMod = sheet.cell(row=MODULE_ROW, column=4).value
        period = sheet.cell(row=MODULE_ROW, column=6).value
        verif = Module.objects.filter(abbrev=idMod, train_prog__abbrev=tpMod, period__name=period)


        if not verif.exists():

            codeMod = sheet.cell(row=MODULE_ROW, column=2).value
            nameMod = sheet.cell(row=MODULE_ROW, column=3).value
            tpMod = sheet.cell(row=MODULE_ROW, column=4).value
            profMod = sheet.cell(row=MODULE_ROW, column=5).value
            tpModule = TrainingProgramme.objects.get(abbrev=tpMod)
            profesMod = Tutor.objects.get(username=profMod)
            periodMod = Period.objects.get(name=period, department=department)

            try:

                module = Module(nom=nameMod, abbrev=idMod, ppn=codeMod, train_prog=tpModule, head=profesMod, period=periodMod)
                module.save()

            except IntegrityError as ie:
                print("A constraint has not been respected creating the Module %s : \n" % idMod, ie)
                pass

        MODULE_ROW+=1

        idMod=sheet.cell(row=MODULE_ROW, column=1).value

    print("Modules extraction done")

def slots_extract(department, book):

    sheet = book["Creneaux"]

    answer = sheet.cell(row=3, column=12).value
    for d in Day.CHOICES[:5]:
        day, created = Day.objects.get_or_create(day=d[0])
    if created:
        day.save()

    if answer == "Oui":
        print("'normal' slots imported")
        default_times=[(8,0),(9,30),(11,0),(14,15),(15,45),(17,15)]
        for hours,minutes in default_times:
            time, created = Time.objects.get_or_create(hours=hours, minutes=minutes)
            if created:
                time.save()
            for day in Day.objects.all():
                slot, created = Slot.objects.get_or_create(duration=90, jour=day, heure=time)
                if created:
                    slot.save()
        assign_day_time_numbers()
        return None

    CRENEAU_ROW=3

    dura = sheet.cell(row=CRENEAU_ROW, column=1).value

    while dura is not None:

        TIME_COL = 3

        day = sheet.cell(row=CRENEAU_ROW, column=2).value
        hour = sheet.cell(row=CRENEAU_ROW, column=TIME_COL).value

        if day == "Tous les jours":
            days = [d[0] for d in Day.CHOICES[:5]]
        elif day == "Lundi":
            days = [Day.MONDAY]
        elif day == "Mardi":
            days = [Day.TUESDAY]
        elif day == "Mercredi":
            days = [Day.WEDNESDAY]
        elif day == "Jeudi":
            days = [Day.THURSDAY]
        elif day == "Vendredi":
            days = [Day.FRIDAY]


        while hour is not None:

            try:

                time = hour.split('h')
                hours = int(time[0])

                if time[1] != "":

                    minutes = int(time[1])

                else:

                    minutes = 0
                slotTime, created = Time.objects.get_or_create(hours=hours, minutes=minutes)

                if created:

                    slotTime.save()

                for day in days:

                    slot, created = Slot.objects.get_or_create(duration=dura,
                                                               jour=Day.objects.get(day=day),
                                                               heure=slotTime)
                    if created:

                        slot.save()

            except IntegrityError as ie:

                print("A constraint has not been respected creating the Slot %s : \n" % dura, ie)
                pass

            TIME_COL += 1
            hour = sheet.cell(row=CRENEAU_ROW, column=TIME_COL).value

        CRENEAU_ROW += 1
        dura = sheet.cell(row=CRENEAU_ROW, column=TIME_COL).value

    assign_day_time_numbers()

    print("Slots extraction done")

def courses_extract(department, book):

    sheet = book['Cours']

    TYPE_ROW = 3

    idType = sheet.cell(row=TYPE_ROW, column=1).value

    while idType is not None:

        TYPE_COL = 2

        verif = CourseType.objects.filter(name=idType, department=department)

        if not verif.exists():

            try:

                course = CourseType(name=idType, department=department)
                course.save()

            except IntegrityError as ie:

                print("A constraint has not been respected creating the CourseType %s : \n" % idType, ie)
                pass

            idGroup = sheet.cell(row=TYPE_ROW, column=TYPE_COL).value

            while idGroup is not None:

                group = GroupType.objects.get(name=idGroup, department=department)
                course.group_types.add(group)
                course.save()

                TYPE_COL += 1
                idGroup = sheet.cell(row=TYPE_ROW, column=TYPE_COL).value

        TYPE_ROW += 1
        idType = sheet.cell(row=TYPE_ROW, column=1).value

    print("Courses' types extraction done")



def displayInfo():

    print("The Professors are : ")

    for p in Tutor.objects.all():

        print(p.username, " : ", p.first_name, " ", p.last_name.upper())

        if p.status == Tutor.SUPP_STAFF :
            print("Qualite : ", p.qualite)
            print("Employer : ", p.employer)
            print("Status : SupplyStaff")
        else:
            print("Status : FullStaff")

        print("e-mail : ", p.email, "\n")

    print("------------------")

    print("The Rooms are : ")

    for r in Room.objects.all():

        print(r.name, ", subroom of : ")

        for sub in r.subroom_of.all():

            print(sub)

    print("---")

    print("The Room groups are : ")

    for rg in RoomGroup.objects.all():

        print(rg.name, ", types : ")

        for t in rg.types.all():

            print("| ", t, " |")

    print("---")

    print("The Room types are : ")

    for rt in RoomType.objects.all():

        print(rt)

    print("------------------")

    print("The Training programs are : ")

    for tp in TrainingProgramme.objects.all():

        print(tp.abbrev, " : ", tp)

    print("---")

    print("The Group types are : ")

    for gt in GroupType.objects.all():

        print(gt)

    print("---")

    print("The Groups are : ")

    for g in Group.objects.all():

        print(g, " : ")
        print(" - size : ", g.size)
        print(" - type : ", g.type)
        print(" - basic : ", g.basic)

        print(" - parent groups : ")
        for i in g.ancestor_groups():
            print(i)

    print("------------------")

    print("The Modules are : ")

    for m in Module.objects.all():

        print(m, " : ", m.nom, " - ", m.ppn)
        print("- head : ", m.head)
        print("- training program : ", m.train_prog)
        print("- starting week : ", m.period.starting_week)
        print("- ending week : ", m.period.ending_week)

    print("------------------")

    print("The Slots are : ")

    for s in Slot.objects.all():

        print(s, " ", s.duration)
